import openpyxl
import json

file_path = r'C:\Users\hp\Downloads\Suivi des Incidents DP 2026 FF 06.05.2026.xlsm'

try:
    print(f"Loading {file_path} in read_only mode...")
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    print(f"Sheets: {wb.sheetnames}")
    
    for sheet_name in wb.sheetnames:
        print(f"\n--- Sheet: {sheet_name} ---")
        sheet = wb[sheet_name]
        count = 0
        for row in sheet.iter_rows(min_row=1, max_row=5, values_only=True):
            print(row)
            count += 1
        if count == 0:
            print("Empty sheet or could not read rows.")
            
except Exception as e:
    print(f"Error: {e}")
