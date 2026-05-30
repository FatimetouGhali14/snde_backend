import openpyxl

file_path = r'C:\Users\hp\Downloads\Suivi des Incidents DP 2026 FF 06.05.2026.xlsm'

try:
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    
    # Try to find any sheet with a lot of data
    for sn in wb.sheetnames:
        sheet = wb[sn]
        # Count non-empty rows
        row_count = 0
        for row in sheet.iter_rows(values_only=True):
            if any(row):
                row_count += 1
            if row_count > 400: # We found a large sheet!
                break
        print(f"Sheet: {sn}, Approx Rows: {row_count}")
            
except Exception as e:
    print(f"Error: {e}")
