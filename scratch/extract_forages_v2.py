import openpyxl

file_path = r'C:\Users\hp\Downloads\Suivi des Incidents DP 2026 FF 06.05.2026.xlsm'

try:
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    
    for sheet_name in ['Nombre de forages par brigade', 'Feuil7', 'Feuil1']:
        if sheet_name in wb.sheetnames:
            print(f"\n--- Sheet: {sheet_name} ---")
            sheet = wb[sheet_name]
            for row in sheet.iter_rows(min_row=1, max_row=20, values_only=True):
                print(row)
                
except Exception as e:
    print(f"Error: {e}")
