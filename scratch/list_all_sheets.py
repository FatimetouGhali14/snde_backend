import openpyxl

file_path = r'C:\Users\hp\Downloads\Suivi des Incidents DP 2026 FF 06.05.2026.xlsm'

try:
    wb = openpyxl.load_workbook(file_path)
    print(f"All sheets: {wb.sheetnames}")
    for sn in wb.sheetnames:
        sheet = wb[sn]
        print(f"Sheet: {sn}, Hidden: {sheet.sheet_state}")
            
except Exception as e:
    print(f"Error: {e}")
