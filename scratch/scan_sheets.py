import openpyxl

file_path = r'C:\Users\hp\Downloads\Suivi des Incidents DP 2026 FF 06.05.2026.xlsm'

try:
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    print(f"Sheets: {wb.sheetnames}")
    
    for sn in wb.sheetnames:
        print(f"\n--- Checking {sn} ---")
        sheet = wb[sn]
        # Check first 2 rows and first 10 columns
        for row in sheet.iter_rows(min_row=1, max_row=2, max_col=15, values_only=True):
            print(row)
            
except Exception as e:
    print(f"Error: {e}")
