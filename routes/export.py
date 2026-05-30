import io
from flask import Blueprint, request, jsonify, send_file
from datetime import datetime, timezone
from config.database import get_db
from middleware.auth import role_required

export_bp = Blueprint("export", __name__)


@export_bp.route("/excel", methods=["GET"])
@role_required("directeur", "admin")
def export_excel():
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return jsonify({"error": "openpyxl non installé"}), 500

    db = get_db()
    incidents = list(db.incidents.find({}).sort("date_declaration", -1))

    wb = Workbook()
    ws = wb.active
    ws.title = "Suivi des incidents"

    headers = [
        "Date de déclaration", "Site", "Localisation",
        "Description de la défaillance", "Impact sur la production",
        "Statut de l'intervention", "Action Corrective", "Brigade",
        "Chef de brigade", "Pièces de rechange", "Date de clôture",
        "Code GMAO", "Observation"
    ]

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", size=10)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    for row_idx, inc in enumerate(incidents, 2):
        def fmt_date(d):
            if isinstance(d, datetime):
                return d.strftime("%d/%m/%Y")
            return ""

        row_data = [
            fmt_date(inc.get("date_declaration")),
            inc.get("site", ""),
            inc.get("localisation", ""),
            inc.get("description", ""),
            inc.get("impact", ""),
            inc.get("statut", ""),
            inc.get("action_corrective", ""),
            inc.get("brigade", ""),
            inc.get("chef_brigade", ""),
            inc.get("pieces_rechange", ""),
            fmt_date(inc.get("date_cloture")),
            inc.get("code_gmao", ""),
            inc.get("observation", "")
        ]
        for col, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col, value=value)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"SNDE_Incidents_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename
    )


@export_bp.route("/import-excel", methods=["POST"])
@role_required("admin")
def import_excel():
    try:
        import pandas as pd
    except ImportError:
        return jsonify({"error": "pandas non installé"}), 500

    keys = list(request.files.keys())
    if not keys:
        return jsonify({"error": "Aucun fichier reçu"}), 400

    f = request.files[keys[0]]

    if not f.filename.endswith((".xlsx", ".xlsm", ".xls")):
        return jsonify({"error": "Format accepté : .xlsx, .xlsm"}), 400

    try:
        df = pd.read_excel(f, sheet_name="Suivi des incidents", header=1)
    except Exception as e:
        return jsonify({"error": f"Impossible de lire le fichier : {str(e)}"}), 400

    # Détection automatique des colonnes
    col_desc    = next((c for c in df.columns if "Description" in str(c)), None)
    col_impact  = next((c for c in df.columns if "Impacte" in str(c)), None)
    col_statut  = next((c for c in df.columns if "Statut" in str(c)), None)
    col_date    = next((c for c in df.columns if "declaration" in str(c)), None)
    col_cloture = next((c for c in df.columns if "cloture" in str(c)), None)
    col_chef    = next((c for c in df.columns if "chef" in str(c).lower()), None)
    col_pieces  = next((c for c in df.columns if "rechange" in str(c).lower()), None)

    db = get_db()
    inseres = 0
    erreurs = []
    now = datetime.now(timezone.utc)

    def parse_date(val):
        try:
            if pd.isna(val):
                return None
        except Exception:
            pass
        if isinstance(val, datetime):
            return val.replace(tzinfo=timezone.utc)
        try:
            return pd.to_datetime(val).to_pydatetime().replace(tzinfo=timezone.utc)
        except Exception:
            return None

    for idx, row in df.iterrows():
        try:
            site = str(row.get("Site", "")).strip()
            desc = str(row.get(col_desc, "")).strip() if col_desc else ""

            if not site or not desc or site == "nan" or desc == "nan":
                continue

            statut_raw = str(row.get(col_statut, "En attente")).strip() if col_statut else "En attente"
            statut = statut_raw if statut_raw in [
                "Acheve", "En attente", "Abandonne", "Debloque apres 4h"
            ] else "En attente"

            impact_raw = str(row.get(col_impact, "Faible")).strip() if col_impact else "Faible"
            impact = impact_raw if impact_raw in [
                "Critique", "Majeure", "Moyenne", "Mineure"
            ] else "Mineure"

            incident = {
                "date_declaration": parse_date(row.get(col_date)) or now,
                "site": site,
                "localisation": str(row.get("Localisation", "")).strip(),
                "description": desc,
                "impact": impact,
                "statut": statut,
                "action_corrective": str(row.get("Action Corrective", "")).strip(),
                "brigade": str(row.get("Brigade", "")).strip(),
                "chef_brigade": str(row.get(col_chef, "")).strip() if col_chef else "",
                "pieces_rechange": str(row.get(col_pieces, "")).strip() if col_pieces else "",
                "date_cloture": parse_date(row.get(col_cloture)) if col_cloture else None,
                "code_gmao": str(row.get("Code GMAO", "")).strip(),
                "observation": str(row.get("Observation", "")).strip(),
                "photo_url": "",
                "source": "import_excel",
                "created_at": now,
                "updated_at": now
            }

            for k, v in incident.items():
                if isinstance(v, str) and v.lower() in ("nan", "none", ""):
                    incident[k] = ""

            db.incidents.insert_one(incident)
            inseres += 1

        except Exception as e:
            erreurs.append(f"Ligne {idx + 2} : {str(e)}")

    return jsonify({
        "message": f"{inseres} incidents importés avec succès",
        "inseres": inseres,
        "erreurs": erreurs[:20]
    }), 200